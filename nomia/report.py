"""Generates the Excel decision log: one workbook per applied run, with a row for every file
the pipeline ever looked at - not just the ones that moved. This is the "explainable" half of
Nomia's core engineering story (see CLAUDE.md / README): every decision has a reason, and every
run's integrity is independently checkable in the Verification sheet.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from nomia.organizer import ItemRecord, RunRecord, UndoJournal

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="FF2C2C2E", end_color="FF2C2C2E", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_AUTO_FILL = PatternFill(start_color="FFDDF3E4", end_color="FFDDF3E4", fill_type="solid")
_REVIEW_FILL = PatternFill(start_color="FFFCEFD0", end_color="FFFCEFD0", fill_type="solid")
_PROBLEM_FILL = PatternFill(start_color="FFFAE1E1", end_color="FFFAE1E1", fill_type="solid")
_NEUTRAL_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

_ROUTE_FILLS = {
    "auto": _AUTO_FILL,
    "review": _REVIEW_FILL,
    "unsorted": _PROBLEM_FILL,
    "other": _NEUTRAL_FILL,
    "left_untouched": _NEUTRAL_FILL,
    "skip_duplicate": _NEUTRAL_FILL,
    "skip_already_organized": _NEUTRAL_FILL,
}

_DECISIONS_COLUMNS = [
    ("Source Path", "source_path", 45),
    ("Status", "status", 12),
    ("Route", "route", 20),
    ("Destination Path", "dest_path", 45),
    ("Dump Copy Path", "dump_path", 40),
    ("Source Preserved", "source_preserved_label", 14),
    ("Category", "category", 16),
    ("Subcategory", "subcategory", 16),
    ("Description", "description", 28),
    ("Confidence", "confidence", 11),
    ("Reason", "reason", 40),
    ("Naming Index", "naming_index", 11),
    ("Date Source", "naming_date_source", 15),
    ("Duplicate Of", "duplicate_of", 40),
    ("Error / Note", "error_label", 30),
    ("Applied At", "applied_at", 22),
]


def _row_values(item: ItemRecord) -> dict:
    return {
        "source_path": item.source_path,
        "status": item.status,
        "route": item.route,
        "dest_path": item.dest_path or "",
        "dump_path": item.dump_path or "",
        "source_preserved_label": "yes" if item.source_preserved else "",
        "category": item.category or "",
        "subcategory": item.subcategory or "",
        "description": item.description or "",
        "confidence": round(item.confidence, 2) if item.confidence is not None else "",
        "reason": item.reason or "",
        "naming_index": item.naming_index if item.naming_index is not None else "",
        "naming_date_source": item.naming_date_source or "",
        "duplicate_of": item.duplicate_of or "",
        "error_label": item.warning or item.error or "",
        "applied_at": item.applied_at or "",
    }


def _write_decisions_sheet(ws: Worksheet, items: list[ItemRecord]) -> None:
    ws.title = "Decisions"
    headers = [label for label, _key, _width in _DECISIONS_COLUMNS]
    ws.append(headers)
    for col_idx, (_label, _key, width) in enumerate(_DECISIONS_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"

    for item in items:
        values = _row_values(item)
        row = [values[key] for _label, key, _width in _DECISIONS_COLUMNS]
        ws.append(row)
        row_idx = ws.max_row
        fill = _ROUTE_FILLS.get(item.route, _NEUTRAL_FILL)
        if item.status == "failed":
            fill = _PROBLEM_FILL
        for col_idx in range(1, len(_DECISIONS_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx, (_label, key, _width) in enumerate(_DECISIONS_COLUMNS, start=1):
        if key in ("reason", "description", "dest_path", "dump_path"):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=False, vertical="top")


def _write_summary_sheet(ws: Worksheet, run: RunRecord, items: list[ItemRecord]) -> None:
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    def _section(title: str) -> None:
        row = ws.max_row + 1 if ws.max_row > 1 else 1
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=13)
        ws.append([])

    def _kv(label: str, value: object) -> None:
        ws.append([label, value])

    _section("Nomia Run Summary")
    _kv("Run ID", run.run_id)
    _kv("Status", run.status)
    _kv("Started At", run.started_at)
    _kv("Finished At", run.finished_at or "")
    _kv("Model Used", run.model_used)
    _kv("Naming Template", run.naming_template)
    _kv("Source Folders", ", ".join(run.source_folders))
    _kv("Destination Root", run.destination_root)

    ws.append([])
    _section("Counts by Status")
    counts_by_status: dict[str, int] = {}
    for item in items:
        counts_by_status[item.status] = counts_by_status.get(item.status, 0) + 1
    for status, count in sorted(counts_by_status.items()):
        _kv(status, count)

    ws.append([])
    _section("Counts by Route")
    counts_by_route: dict[str, int] = {}
    for item in items:
        counts_by_route[item.route] = counts_by_route.get(item.route, 0) + 1
    for route, count in sorted(counts_by_route.items()):
        _kv(route, count)

    ws.append([])
    _section("Counts by Category (applied only)")
    counts_by_category: dict[str, int] = {}
    for item in items:
        if item.status == "applied" and item.category:
            counts_by_category[item.category] = counts_by_category.get(item.category, 0) + 1
    for category, count in sorted(counts_by_category.items()):
        _kv(category, count)


def _write_verification_sheet(ws: Worksheet, run: RunRecord) -> None:
    ws.title = "Verification"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    verification = run.verification_json
    if not verification:
        ws.append(["No verification data recorded for this run."])
        return

    ok = verification["scanned_count"] == verification["accounted_count"] and not verification["hash_mismatches"]
    header_cell = ws.cell(row=1, column=1, value="ALL CHECKS PASSED" if ok else "VERIFICATION ISSUES FOUND")
    header_cell.font = Font(bold=True, size=13, color="FF1B7A3D" if ok else "FFB3261E")
    ws.append([])

    ws.append(["Scanned Count", verification["scanned_count"]])
    ws.append(["Accounted For Count", verification["accounted_count"]])
    ws.append(["Counts Match", "yes" if verification["scanned_count"] == verification["accounted_count"] else "NO - see log"])
    ws.append([])
    ws.append(["Hash Matches", verification["hash_matches"]])
    ws.append(["Hash Mismatches", len(verification["hash_mismatches"])])
    ws.append([])

    if verification["counts_by_status"]:
        ws.append(["Counts by Status"])
        for status, count in sorted(verification["counts_by_status"].items()):
            ws.append([status, count])
        ws.append([])

    if verification["hash_mismatches"]:
        ws.append(["Hash Mismatch Details"])
        mismatch_headers = ["Item ID", "Source Path", "Kind", "Expected", "Actual"]
        ws.append(mismatch_headers)
        for col_idx in range(1, len(mismatch_headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for mismatch in verification["hash_mismatches"]:
            ws.append([
                mismatch["item_id"], mismatch["source_path"], mismatch["kind"],
                mismatch["expected"], mismatch.get("actual") or "",
            ])
            for col_idx in range(1, len(mismatch_headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = _PROBLEM_FILL


def generate_report(run_id: str, journal: UndoJournal, out_path: Path) -> Path:
    run = journal.get_run(run_id)
    if run is None:
        raise ValueError(f"No such run: {run_id}")
    items = journal.get_items(run_id)

    workbook = Workbook()
    _write_decisions_sheet(workbook.active, items)
    _write_summary_sheet(workbook.create_sheet(), run, items)
    _write_verification_sheet(workbook.create_sheet(), run)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    journal.set_run_report_path(run_id, str(out_path))
    logger.info("Wrote report for run %s to %s", run_id, out_path)
    return out_path
