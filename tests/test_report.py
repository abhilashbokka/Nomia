import hashlib
from pathlib import Path

from openpyxl import load_workbook

from nomia.config import NomiaConfig
from nomia.organizer import UndoJournal, apply_plan
from nomia.report import generate_report


def _hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _make_source(tmp_path: Path, name: str, content: bytes = b"file-content") -> Path:
    src_dir = tmp_path / "source"
    src_dir.mkdir(exist_ok=True)
    path = src_dir / name
    path.write_bytes(content)
    return path


def _plan_and_apply(tmp_path: Path, journal: UndoJournal) -> str:
    content = b"file-content"
    source = _make_source(tmp_path, "receipt.pdf", content)
    dest_root = tmp_path / "dest"
    run_id = journal.start_run(
        source_folders=[str(tmp_path / "source")], destination_root=str(dest_root),
        naming_template="{category}_{yyyy}-{mm}-{dd}_{index}", model_used="moondream",
        thresholds={"auto_min": 0.8, "review_min": 0.5}, config_snapshot={},
    )
    journal.record_planned(
        run_id, source_path=str(source), source_sha256=_hash(content), source_size=len(content),
        dest_path="receipt/receipt_2024-01-01.pdf", dump_path="_dump/receipt.pdf", route="auto",
        category="receipt", subcategory="grocery", description="costco-receipt", confidence=0.91,
        reason="Looks like a receipt",
    )
    journal.set_run_totals(run_id, {"scanned_count": 1})
    apply_plan(run_id, NomiaConfig(), journal)
    return run_id


def test_generate_report_writes_all_three_sheets(tmp_path):
    journal = UndoJournal(tmp_path / "journal.sqlite3")
    run_id = _plan_and_apply(tmp_path, journal)

    out_path = tmp_path / "report.xlsx"
    result_path = generate_report(run_id, journal, out_path)

    assert result_path == out_path
    assert out_path.exists()

    wb = load_workbook(out_path)
    assert wb.sheetnames == ["Decisions", "Summary", "Verification"]


def test_decisions_sheet_has_header_and_one_data_row(tmp_path):
    journal = UndoJournal(tmp_path / "journal.sqlite3")
    run_id = _plan_and_apply(tmp_path, journal)
    out_path = tmp_path / "report.xlsx"
    generate_report(run_id, journal, out_path)

    wb = load_workbook(out_path)
    ws = wb["Decisions"]
    assert ws.cell(row=1, column=1).value == "Source Path"
    assert ws.max_row == 2  # header + 1 item
    values = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert "receipt.pdf" in str(values[0])
    assert values[1] == "applied"
    assert values[2] == "auto"


def test_verification_sheet_reports_clean_run(tmp_path):
    journal = UndoJournal(tmp_path / "journal.sqlite3")
    run_id = _plan_and_apply(tmp_path, journal)
    out_path = tmp_path / "report.xlsx"
    generate_report(run_id, journal, out_path)

    wb = load_workbook(out_path)
    ws = wb["Verification"]
    assert ws.cell(row=1, column=1).value == "ALL CHECKS PASSED"


def test_verification_sheet_reports_problems_after_corruption(tmp_path):
    journal = UndoJournal(tmp_path / "journal.sqlite3")
    run_id = _plan_and_apply(tmp_path, journal)

    dest_file = tmp_path / "dest" / "receipt" / "receipt_2024-01-01.pdf"
    dest_file.write_bytes(b"corrupted")
    from nomia.organizer import verify_run
    verify_run(run_id, journal)

    out_path = tmp_path / "report.xlsx"
    generate_report(run_id, journal, out_path)

    wb = load_workbook(out_path)
    ws = wb["Verification"]
    assert ws.cell(row=1, column=1).value == "VERIFICATION ISSUES FOUND"


def test_summary_sheet_includes_run_metadata_and_counts(tmp_path):
    journal = UndoJournal(tmp_path / "journal.sqlite3")
    run_id = _plan_and_apply(tmp_path, journal)
    out_path = tmp_path / "report.xlsx"
    generate_report(run_id, journal, out_path)

    wb = load_workbook(out_path)
    ws = wb["Summary"]
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert run_id in all_values
    assert "moondream" in all_values
    assert "receipt" in all_values  # category count section


def test_generate_report_sets_run_report_path(tmp_path):
    journal = UndoJournal(tmp_path / "journal.sqlite3")
    run_id = _plan_and_apply(tmp_path, journal)
    out_path = tmp_path / "reports" / "report.xlsx"
    generate_report(run_id, journal, out_path)

    run = journal.get_run(run_id)
    assert run.report_path == str(out_path)
